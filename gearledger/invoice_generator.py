# -*- coding: utf-8 -*-
"""
Generate invoices from results file using catalog data.
"""
import os
import pandas as pd
from datetime import datetime
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def generate_invoice_from_results(
    results_path: str, catalog_path: str, output_path: str = None
) -> Dict[str, Any]:
    """
    Generate an invoice file from the results ledger.

    Args:
        results_path: Path to results Excel file (with Артикул, Клиент, Количество, Вес columns)
        catalog_path: Path to catalog Excel file (with full product info including prices)
        output_path: Path for output invoice file (optional, auto-generated if not provided)

    Returns:
        Dict with 'ok', 'path', 'error' keys
    """
    try:
        # Read results file
        if not os.path.exists(results_path):
            return {
                "ok": False,
                "error": f"Results file not found: {results_path}",
                "path": "",
            }

        results_df = pd.read_excel(results_path)

        # Read catalog file
        if not os.path.exists(catalog_path):
            return {
                "ok": False,
                "error": f"Catalog file not found: {catalog_path}",
                "path": "",
            }

        catalog_df = pd.read_excel(catalog_path)

        # Detect column names in catalog (they might vary)
        col_mapping = _detect_catalog_columns(catalog_df)

        if not col_mapping.get("номер"):
            return {
                "ok": False,
                "error": "Could not find part number column in catalog",
                "path": "",
            }

        # Generate output path if not provided
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_dir = os.path.dirname(results_path) or "."
            output_path = os.path.join(output_dir, f"invoice_{timestamp}.xlsx")

        # Create invoice workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice"

        # Group results by client
        if "Клиент" not in results_df.columns:
            return {
                "ok": False,
                "error": "Results file missing 'Клиент' column",
                "path": "",
            }

        clients = results_df["Клиент"].unique()

        current_row = 1

        for client in clients:
            if pd.isna(client) or str(client).strip() == "":
                continue

            client_items = results_df[results_df["Клиент"] == client]

            # Write client header
            current_row = _write_client_section(
                ws, client, client_items, catalog_df, col_mapping, current_row
            )

            current_row += 2  # Add spacing between clients

        # Save workbook
        wb.save(output_path)

        return {"ok": True, "path": output_path, "error": "", "clients": len(clients)}

    except Exception as e:
        return {"ok": False, "error": str(e), "path": ""}


def _detect_catalog_columns(df: pd.DataFrame) -> Dict[str, str]:
    """Detect column names in catalog DataFrame."""
    mapping = {}

    for col in df.columns:
        col_lower = str(col).strip().lower()

        # Detect part number column
        if "номер" in col_lower or "artikul" in col_lower or "арт" in col_lower:
            mapping["номер"] = col

        # Detect brand column
        if "бренд" in col_lower or "brand" in col_lower or "брэнд" in col_lower:
            mapping["бренд"] = col

        # Detect description column
        if (
            "описание" in col_lower
            or "description" in col_lower
            or "наименование" in col_lower
        ):
            mapping["описание"] = col

        # Detect price column
        if "цена продажи" in col_lower or "price" in col_lower or "цена" in col_lower:
            if "продажи" in col_lower or "sale" in col_lower:
                mapping["цена"] = col

        # Detect client column
        if "клиент" in col_lower or "client" in col_lower:
            mapping["клиент"] = col

    # Fallback to exact names (prioritize Номер over Артикул for new format)
    if "Номер" in df.columns:
        mapping["номер"] = "Номер"
    elif "Артикул" in df.columns:
        mapping["номер"] = "Артикул"
    if "Брэнд" in df.columns:
        mapping["бренд"] = "Брэнд"
    if "Описание" in df.columns:
        mapping["описание"] = "Описание"
    if "Цена продажи" in df.columns:
        mapping["цена"] = "Цена продажи"

    return mapping


def _write_client_section(
    ws,
    client: str,
    items: pd.DataFrame,
    catalog_df: pd.DataFrame,
    col_mapping: Dict[str, str],
    start_row: int,
) -> int:
    """Write a section for one client. Returns the next available row."""

    current_row = start_row

    # Header: Покупатель
    ws.merge_cells(f"A{current_row}:L{current_row}")
    cell = ws[f"A{current_row}"]
    cell.value = f"Покупатель: {client}"
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    current_row += 2

    # Date header
    ws.merge_cells(f"B{current_row}:D{current_row}")
    cell = ws[f"B{current_row}"]
    cell.value = "Дата"
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f"E{current_row}:G{current_row}")
    cell = ws[f"E{current_row}"]
    cell.value = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    current_row += 1

    # Table headers
    headers = [
        "№",
        "Брэнд",
        "Номер",
        "Описание",
        "Кол.",
        "Цена продажи",
        "Сумма продажи",
    ]
    header_cols = ["A", "B", "C", "D", "E", "F", "G"]

    for col, header in zip(header_cols, headers):
        cell = ws[f"{col}{current_row}"]
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(
            start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
        )

        # Add borders
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        cell.border = thin_border

    current_row += 1
    data_start_row = current_row

    # Write item rows
    item_num = 1
    total_sum = 0

    for _, result_row in items.iterrows():
        artikul = str(result_row.get("Артикул", ""))
        quantity = result_row.get("Количество", 1)

        # Look up in catalog
        item_data = _lookup_in_catalog(artikul, catalog_df, col_mapping)

        if item_data:
            brand = item_data.get("бренд", "")
            number = item_data.get("номер", artikul)
            description = item_data.get("описание", "")
            price = item_data.get("цена", 0)

            try:
                price = float(price) if price else 0
                quantity = int(quantity) if quantity else 1
            except:
                price = 0
                quantity = 1

            total = price * quantity
            total_sum += total

            # Write row
            ws[f"A{current_row}"] = item_num
            ws[f"B{current_row}"] = brand
            ws[f"C{current_row}"] = number
            ws[f"D{current_row}"] = description
            ws[f"E{current_row}"] = quantity
            ws[f"F{current_row}"] = f"{price:.2f}"
            ws[f"G{current_row}"] = f"{total:.2f}"

            # Format cells
            for col in header_cols:
                cell = ws[f"{col}{current_row}"]
                cell.alignment = Alignment(
                    horizontal="center" if col in ["A", "E"] else "left",
                    vertical="center",
                )
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

            item_num += 1
            current_row += 1

    # Write total row
    ws[f"D{current_row}"] = "Итого:"
    ws[f"D{current_row}"].font = Font(bold=True)
    ws[f"D{current_row}"].alignment = Alignment(horizontal="right", vertical="center")

    ws[f"E{current_row}"] = item_num - 1
    ws[f"E{current_row}"].font = Font(bold=True)
    ws[f"E{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

    ws[f"F{current_row}"] = f"{total_sum:.2f}"
    ws[f"F{current_row}"].font = Font(bold=True)
    ws[f"F{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

    ws[f"G{current_row}"] = f"{total_sum:.2f}"
    ws[f"G{current_row}"].font = Font(bold=True)
    ws[f"G{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

    # Add borders to total row
    for col in ["D", "E", "F", "G"]:
        ws[f"{col}{current_row}"].border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="double"),
        )
        ws[f"{col}{current_row}"].fill = PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
        )

    current_row += 1

    # Set column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15

    return current_row


def _lookup_in_catalog(
    artikul: str, catalog_df: pd.DataFrame, col_mapping: Dict[str, str]
) -> Dict[str, Any]:
    """Look up item details in catalog by artikul."""

    def normalize(s: str) -> str:
        return str(s or "").replace(" ", "").replace("-", "").replace(".", "").upper()

    artikul_norm = normalize(artikul)

    # Find matching row in catalog
    номер_col = col_mapping.get("номер")
    if not номер_col:
        return {}

    # Create normalized column for matching
    catalog_df["_NORM_TEMP"] = catalog_df[номер_col].astype(str).apply(normalize)

    matches = catalog_df[catalog_df["_NORM_TEMP"] == artikul_norm]

    if matches.empty:
        return {}

    row = matches.iloc[0]

    result = {}
    if col_mapping.get("бренд"):
        result["бренд"] = row.get(col_mapping["бренд"], "")
    if col_mapping.get("номер"):
        result["номер"] = row.get(col_mapping["номер"], "")
    if col_mapping.get("описание"):
        result["описание"] = row.get(col_mapping["описание"], "")
    if col_mapping.get("цена"):
        result["цена"] = row.get(col_mapping["цена"], 0)

    return result
